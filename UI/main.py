import customtkinter as ctk
import tkinter as tk
import openpyxl
import re

# Python Type Hint
from typing import List

# Provides the Selenium WebDriver API for browser automation.
from selenium import webdriver

# Offers various methods for identifying web elements in Selenium.
from selenium.webdriver.common.by import By

# ChromeOptions
options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)

# Chrome WebDriver
driver = webdriver.Chrome(options=options)
driver.get("https://www.amazon.com/")

from selenium.common.exceptions import NoSuchElementException

import time
import os
import requests
from PIL import Image, ImageOps, ImageTk
import pytesseract
import shutil
import requests
from bs4 import BeautifulSoup
from googletrans import Translator


class MainUI:
    def __init__(self, font, font_size):
        self.font = font
        self.font_size = font_size

        self.manageVaribles = ManageVariables()
        self.excelCRUD = ExcelCRUD()

        # init UI
        self.UI = False  # flag
        self.id_entry = None
        self.password_entry = None
        self.naver_checkbox = None
        self.coupang_checkbox = None
        self.url_entry = None
        self.url_scrollable_frame = None
        self.log_textbox = None
        self.product_image_canvas = None
        self.suspicious_ingredients_textbox = None
        self.code_lable = None
        self.warning_lable = None
        self.scrollable_frame = None

        self.naver_thumbnail_size = 1000
        self.naver_image_size = 860
        self.coupangr_thumbnail_size = 500
        self.coupang_image_size = 780

    def start(self):
        root = ctk.CTk()
        root.title("KME Scraper")

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.font_style = ctk.CTkFont(self.font, size=self.font_size)

        """
        frame 01
        """
        frame_01 = ctk.CTkFrame(root)
        frame_01.grid(row=0, column=0, padx=(10, 5), pady=(10, 5), sticky="new")

        option_frame = ctk.CTkFrame(frame_01)
        option_frame.pack(side="left", fill="both", expand=True, padx=(10, 5), pady=10)

        amazon_iherb_segemented_button = ctk.CTkSegmentedButton(
            option_frame,
            font=self.font_style,
            values=[
                " 아마존 ",
                " 아이허브 ",
            ],
            command=self.manageVaribles.update_amazon_iherb_option,
        )
        amazon_iherb_segemented_button.set(" 아마존 ")
        amazon_iherb_segemented_button.pack(expand=True, padx=10, pady=10)

        entry_frame = ctk.CTkFrame(frame_01)
        entry_frame.pack(fill="both", side="right", expand=True, padx=(5, 10), pady=10)

        # Mobile phone number or email = entry
        self.id_entry = ctk.CTkEntry(
            entry_frame,
            placeholder_text=" Mobile phone number or email",
            font=self.font_style,
        )
        self.id_entry.pack(side="top", fill="x", expand=True, padx=20, pady=(10, 2.5))
        # Password = entry
        self.password_entry = ctk.CTkEntry(
            entry_frame, placeholder_text=" Password", font=self.font_style
        )
        self.password_entry.pack(
            side="bottom", fill="x", expand=True, padx=20, pady=(2.5, 10)
        )

        """
        frame 02
        """
        frame_02 = ctk.CTkFrame(root)
        frame_02.grid(row=1, column=0, padx=(10, 5), pady=5, sticky="new")

        image_frame = ctk.CTkFrame(frame_02)
        image_frame.pack(side="top", fill="both", expand=True, padx=10, pady=(10, 2.5))

        # 네이버 = checkbox
        self.naver_checkbox = ctk.CTkCheckBox(
            image_frame, text=" 네이버 (1000 , 860)", font=self.font_style
        )
        self.naver_checkbox.pack(expand=True, side="left", padx=(50, 10), pady=10)
        # 쿠팡 = checkbox
        self.coupang_checkbox = ctk.CTkCheckBox(
            image_frame, text=" 쿠팡 (500 , 780)", font=self.font_style
        )
        self.coupang_checkbox.pack(expand=True, side="left", padx=(10, 50), pady=10)

        border_frame = ctk.CTkFrame(frame_02)
        border_frame.pack(side="top", fill="both", expand=True, padx=10, pady=(2.5, 10))

        # 색상 결정 = segemented button
        color_segemented_button = ctk.CTkSegmentedButton(
            border_frame,
            font=self.font_style,
            values=[
                " White ",
                " Red ",
                " Orange ",
                " Yellow ",
                " Green ",
                " Blue ",
                " Purple ",
            ],
            command=self.manageVaribles.update_thumbnail_border_color,
        )
        color_segemented_button.set(" White ")
        color_segemented_button.pack(expand=True, padx=10, pady=10)

        """
        frame 03
        """
        frame_03 = ctk.CTkFrame(root)
        frame_03.grid(
            row=0, rowspan=4, column=1, padx=(5, 10), pady=(10, 5), sticky="news"
        )
        # 제품 주소 = entry
        self.url_entry = ctk.CTkEntry(
            frame_03, placeholder_text=" 제품 주소", font=self.font_style
        )
        self.url_entry.pack(fill="x", padx=10, pady=(10, 2.5))

        self.url_scrollable_frame = ctk.CTkScrollableFrame(frame_03, width=500)
        self.url_scrollable_frame.pack(fill="both", expand=True, padx=10, pady=2.5)

        # 제품 주소(URL) 중복 검사 => 목록 추가 = button
        add_button = ctk.CTkButton(
            frame_03,
            width=200,
            text="제품 주소 중복 검사 => 목록 추가",
            font=self.font_style,
            command=self.check_duplicates_and_add_URL,
        )
        add_button.pack(fill="x", padx=10, pady=2.5)

        # 제품 사진 수집 => 금지 성분 조사 = button
        add_button = ctk.CTkButton(
            frame_03,
            width=200,
            text="제품 사진 수집 => 금지 성분 조사",
            font=self.font_style,
            command=self.images_and_ingredients,
        )
        add_button.pack(fill="x", padx=10, pady=(2.5, 10))

        """
        text frame
        """
        textbox_frame = ctk.CTkFrame(root)
        textbox_frame.grid(row=2, column=0, padx=(10, 5), pady=5, sticky="news")

        # 설명 출력창 = textbox
        self.log_textbox = ctk.CTkTextbox(
            textbox_frame, height=30, font=self.font_style
        )
        self.log_textbox.pack(fill="both", expand=True, padx=10, pady=10)

        """
        frame 04
        """
        frame_04 = ctk.CTkFrame(root)
        frame_04.grid(row=4, column=0, padx=(10, 5), pady=5, sticky="news")

        frame_04.grid_columnconfigure((1), weight=1)

        # 제품 사진 출력창 = canvas
        self.product_image_canvas = tk.Canvas(frame_04, width=200, height=200)
        self.product_image_canvas.grid(row=0, column=0, padx=(10, 5), pady=(10, 5))

        suspicious_ingredients_textbox_frame = ctk.CTkFrame(frame_04)
        suspicious_ingredients_textbox_frame.grid(
            row=0, column=1, padx=(5, 10), pady=(10, 5), sticky="news"
        )

        # 금지 성분 출력창 = textbox
        self.suspicious_ingredients_textbox = ctk.CTkTextbox(
            suspicious_ingredients_textbox_frame,
            activate_scrollbars=False,
            font=self.font_style,
        )
        self.suspicious_ingredients_textbox.pack(
            side="left", expand=True, fill="both", padx=(5, 0), pady=5
        )

        ctk_textbox_scrollbar = ctk.CTkScrollbar(
            suspicious_ingredients_textbox_frame,
            command=self.suspicious_ingredients_textbox.yview,
        )
        ctk_textbox_scrollbar.pack(side="right", fill="y")

        self.suspicious_ingredients_textbox.configure(
            yscrollcommand=ctk_textbox_scrollbar.set
        )

        product_code = ctk.CTkFrame(frame_04, height=50)
        product_code.grid(row=1, column=0, padx=(10, 5), pady=(5, 10), sticky="ew")

        # 제품 식별자 출력창 = label
        self.code_lable = ctk.CTkLabel(product_code, text="", font=self.font_style)
        self.code_lable.pack(padx=5)

        pass_fail_frame = ctk.CTkFrame(frame_04, height=50)
        pass_fail_frame.grid(row=1, column=1, padx=(5, 10), pady=(5, 10), sticky="ew")

        # 금지 성분 존재 여부 출력창 = label
        self.warning_lable = ctk.CTkLabel(
            pass_fail_frame, text="", width=250, font=self.font_style
        )
        self.warning_lable.pack(padx=5)

        """
        frame 05
        """
        frame_05 = ctk.CTkFrame(root)
        frame_05.grid(row=3, column=0, padx=(10, 5), pady=5, sticky="news")

        license_frame = ctk.CTkFrame(frame_05)
        license_frame.pack(side="left", fill="x", expand=True, padx=(10, 5), pady=10)

        # 사용자 정보 = label
        mac_license_lable = ctk.CTkLabel(
            license_frame, text="40-B0-76-42-8F-**", font=self.font_style
        )
        mac_license_lable.pack(padx=5)

        list_date_frame = ctk.CTkFrame(frame_05)
        list_date_frame.pack(side="left", fill="x", expand=True, padx=5, pady=10)

        # 금지 성분 최신화 = label
        list_date_lable = ctk.CTkLabel(
            list_date_frame, text="금지 성분 최신화. 2022-10-16", font=self.font_style
        )
        list_date_lable.pack(padx=5)

        version_frame = ctk.CTkFrame(frame_05)
        version_frame.pack(side="left", fill="x", expand=True, padx=(5, 10), pady=10)

        # KME. 24.1.0 = label
        version_lable = ctk.CTkLabel(
            version_frame, text="KME 24.1.0", font=self.font_style
        )
        version_lable.pack(padx=5)

        """
        frame 06
        """
        frame_06 = ctk.CTkFrame(root)
        frame_06.grid(row=4, column=1, padx=(5, 10), pady=5, sticky="news")

        self.scrollable_frame = ctk.CTkScrollableFrame(frame_06, width=500)
        self.scrollable_frame.pack(fill="both", expand=True, padx=10, pady=10)

        """
        frame 07
        """
        frame_07 = ctk.CTkFrame(root)
        frame_07.grid(
            row=5, column=0, columnspan=2, padx=10, pady=(5, 10), sticky="news"
        )

        # 제품 번호 = entry
        code_entry = ctk.CTkEntry(
            frame_07, placeholder_text=" 제품 번호", font=self.font_style
        )
        code_entry.pack(fill="x", expand=True, side="left", padx=(10, 5), pady=10)

        # 검색 = button
        search_button = ctk.CTkButton(
            frame_07,
            text="검색",
            font=self.font_style,
            # command=search_product
        )
        search_button.pack(fill="x", expand=True, side="left", padx=5, pady=10)

        # 제품 주소(URL) 출력창 = textbox
        url_textbox = ctk.CTkTextbox(
            frame_07, width=600, height=30, font=self.font_style
        )
        url_textbox.pack(fill="x", expand=True, side="left", padx=5, pady=10)

        # 제품 주소 복사 = button
        copy_url_button = ctk.CTkButton(
            frame_07,
            text="제품 주소 복사",
            font=self.font_style,
            # command=copy_product_url
        )
        copy_url_button.pack(fill="x", expand=True, side="left", padx=(5, 10), pady=10)

        self.UI = True
        root.mainloop()

    def logger(self, message: str):
        if self.UI:
            self.log_textbox.delete("0.0", ctk.END)
            self.log_textbox.insert("0.0", message)

    def check_duplicates_and_add_URL(self):
        option: str = self.manageVaribles.get_amazon_iherb_option()

        url = self.url_entry.get().strip()

        products_urls_from_excel = self.excelCRUD.get_products_urls(option)

        if url in products_urls_from_excel:
            self.logger("[경고] 중복되는 제품 주소가 존재합니다!")
        else:
            self.logger("제품 주소가 추가되었습니다.")

            self.manageVaribles.append_url(url)

            # 제품 주소 = frame
            url_frame = ctk.CTkFrame(self.url_scrollable_frame)
            url_frame.pack(fill="x", pady=(5, 0))

            # 삭제 = button
            delete_button = ctk.CTkButton(
                url_frame,
                text="삭제",
                width=50,
                fg_color="#CC3D3D",
                hover_color="#960707",
                font=self.font_style,
                command=lambda frame=url_frame, url=url: (
                    self.manageVaribles.remove_url(url),
                    self.logger("삭제 완료!"),
                    frame.destroy(),
                ),
            )
            delete_button.pack(side="left", padx=5, pady=5)

            if option == "amazon":
                asin_code = re.search(r"dp\/([A-Z0-9]{10})\/", url).group(1)
                # 제품 번호(asin_code) = button
                asin_code_button = ctk.CTkButton(
                    url_frame,
                    text=f"{asin_code}",
                    width=50,
                    font=self.font_style,
                )
                asin_code_button.pack(side="left", pady=5)
            else:
                product_id = url.rsplit("/", 1)[-1].replace("?rec=home", "")
                # 제품 번호(product_id) = button
                product_id_button = ctk.CTkButton(
                    url_frame, text=f"{product_id}", width=50, font=self.font_style
                )
                product_id_button.pack(side="left", pady=5)

            # 제품 주소 = label
            label = ctk.CTkLabel(url_frame, text=url, font=self.font_style)
            label.pack(side="left", padx=5, pady=5, anchor="center")

    def images_and_ingredients(self):
        option: str = self.manageVaribles.get_amazon_iherb_option()

        self.logger("처리 중입니다. 여유롭게 기다려주세요!")

        if option == "amazon":
            self.login_amazon(1, 0.5)
            self.amazon_crawler(option)
        else:
            self.iherb_crawler(option)

    def login_amazon(self, sleep_time: int, delay_time: int):
        driver.get(
            "https://www.amazon.com/-/ko/ap/signin?openid.pape.max_auth_age=0&openid.return_to=https%3A%2F%2Fwww.amazon.com%2F%3Flanguage%3Dko_KR%26ref_%3Dnav_ya_signin&openid.identity=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0%2Fidentifier_select&openid.assoc_handle=usflex&openid.mode=checkid_setup&openid.claimed_id=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0%2Fidentifier_select&openid.ns=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0&"
        )

        time.sleep(sleep_time)
        input_string = self.id_entry.get()
        delay = delay_time
        input_field = driver.find_element(By.CSS_SELECTOR, "#ap_email")
        for character in input_string:
            input_field.send_keys(character)
            time.sleep(delay)
        time.sleep(sleep_time)
        driver.find_element(By.CSS_SELECTOR, "#continue").click()

        time.sleep(sleep_time)
        input_string = self.password_entry.get()
        delay = delay_time
        input_field = driver.find_element(By.CSS_SELECTOR, "#ap_password")
        for character in input_string:
            input_field.send_keys(character)
            time.sleep(delay)
        time.sleep(sleep_time)
        driver.find_element(By.CSS_SELECTOR, "#signInSubmit").click()

    def amazon_crawler(self, option: str):
        for url in self.manageVaribles.get_urls():
            print(url)
            time.sleep(1)
            driver.get(url)

            product_name = driver.find_element(By.CSS_SELECTOR, "#productTitle").text

            pattern = r"dp\/([A-Z0-9]{10})\/"
            asin_code = re.search(pattern, url).group(1)

            folder_path = os.path.join("./amazon", f"{asin_code}")
            os.makedirs(folder_path, exist_ok=True)
            print(f"{asin_code} => folder created!")

            driver.find_element(By.CSS_SELECTOR, "#imgTagWrapperId").click()
            time.sleep(1)

            num = 0
            while True:
                css_selector = f"#ivImage_{num} > div"
                try:
                    driver.find_element(By.CSS_SELECTOR, css_selector).click()

                    time.sleep(2)
                    image_url = driver.find_element(
                        By.CSS_SELECTOR, "#ivLargeImage > img"
                    ).get_attribute("src")

                    response = requests.get(image_url)
                    filename = f"image{num + 1}.jpg"
                    with open(f"./amazon/{asin_code}/{filename}", "wb+") as f:
                        f.write(response.content)
                    print(f"{num + 1} => save complete!")

                    image = Image.open(f"./amazon/{asin_code}/{filename}")

                    if self.naver_checkbox.get():
                        self.retouch_product_image(
                            "naver", option, asin_code, image, filename, num
                        )
                    if self.coupang_checkbox.get():
                        self.retouch_product_image(
                            "coupang", option, asin_code, image, filename, num
                        )

                    num += 1

                except NoSuchElementException:
                    print("no more images exist")
                    break

            if self.naver_checkbox.get():
                self.create_thumbnail_image("naver", option, asin_code)
            if self.coupang_checkbox.get():
                self.create_thumbnail_image("coupang", option, asin_code)

            self.logger("모든 사진이 정상적으로 저장되었습니다.")

            self.amazon_ocr(asin_code)

            self.logger("금지 성분 검사가 완료되었습니다.")

            self.excelCRUD.update_row(
                "amazon",
                asin_code,
                product_name,
                self.manageVaribles.get_suspicious_ingredients(),
                url,
            )

            # 제품 정보 = frame
            product_frame = ctk.CTkFrame(self.scrollable_frame)
            product_frame.pack(fill="x", pady=(5, 0))

            # 제품 번호(asin_code) = button
            id_button = ctk.CTkButton(
                product_frame,
                text=f"{asin_code}",
                width=50,
                font=self.font_style,
                command=lambda ASIN_CODE=asin_code: self.id_button_callback(
                    option, ASIN_CODE
                ),
            )
            id_button.pack(side="left", padx=(5, 0), pady=5)

            # 합격 = button
            pass_button = ctk.CTkButton(
                product_frame,
                text="pass",
                width=50,
                fg_color="#217346",
                hover_color="#005000",
                font=self.font_style,
                command=lambda ASIN_CODE=asin_code, frame=product_frame: self.pass_button_callback(
                    ASIN_CODE, frame
                ),
            )
            pass_button.pack(side="left", padx=5, pady=5)

            # 불합격 = button
            fail_button = ctk.CTkButton(
                product_frame,
                text="fail",
                width=50,
                fg_color="#CC3D3D",
                hover_color="#960707",
                font=self.font_style,
                command=lambda ASIN_CODE=asin_code, frame=product_frame: self.fail_button_callback(
                    option, ASIN_CODE, frame
                ),
            )
            fail_button.pack(side="left", pady=5)

            # 제품명 = label
            label = ctk.CTkLabel(product_frame, text=product_name, font=self.font_style)
            label.pack(side="left", padx=5, pady=5, anchor="center")

        driver.quit()

    def iherb_crawler(self, option: str):
        for url in self.manageVaribles.get_urls():
            print(url)

            product_id = url.rsplit("/", 1)[-1].replace("?rec=home", "")

            folder_path = os.path.join("./iherb", f"{product_id}")
            os.makedirs(folder_path, exist_ok=True)
            print(f"{product_id} => folder created!")

            fake_header = self.manageVaribles.get_fake_header()
            response = requests.get(url, headers=fake_header)
            soup = BeautifulSoup(response.content, "html.parser")

            product_name = soup.find("div", class_="product-summary-title").get_text(
                strip=True
            )
            img_tags = soup.find_all("img", {"data-large-img": True})

            for num, img_tag in enumerate(img_tags):
                img_url = img_tag["data-large-img"]

                response = requests.get(img_url, headers=fake_header)
                filename = f"image{num + 1}.jpg"
                with open(f"./iherb/{product_id}/{filename}", "wb+") as f:
                    f.write(response.content)
                print(f"{num + 1} => save complete!")

                image = Image.open(f"./iherb/{product_id}/{filename}")

                if self.naver_checkbox.get():
                    self.retouch_product_image(
                        "naver", option, product_id, image, filename, num
                    )
                if self.coupang_checkbox.get():
                    self.retouch_product_image(
                        "coupang", option, product_id, image, filename, num
                    )

            if self.naver_checkbox.get():
                self.create_thumbnail_image("naver", option, product_id)
            if self.coupang_checkbox.get():
                self.create_thumbnail_image("coupang", option, product_id)

            self.logger("모든 사진이 정상적으로 저장되었습니다.")

            self.iherb_ocr(url)

            self.logger("금지 성분 검사가 완료되었습니다.")

            self.excelCRUD.update_row(
                "iherb",
                product_id,
                product_name,
                self.manageVaribles.get_suspicious_ingredients(),
                url,
            )

            # 제품 정보 = frame
            product_frame = ctk.CTkFrame(self.scrollable_frame)
            product_frame.pack(fill="x", pady=(5, 0))

            # 제품 번호(product_id) = button
            id_button = ctk.CTkButton(
                product_frame,
                text=f"{product_id}",
                width=50,
                font=self.font_style,
                command=lambda PRODUCT_ID=product_id: self.id_button_callback(
                    option, PRODUCT_ID
                ),
            )
            id_button.pack(side="left", padx=(5, 0), pady=5)

            # 합격 = button
            pass_button = ctk.CTkButton(
                product_frame,
                text="pass",
                width=50,
                fg_color="#217346",
                hover_color="#005000",
                font=self.font_style,
                command=lambda PRODUCT_ID=product_id, frame=product_frame: self.pass_button_callback(
                    PRODUCT_ID, frame
                ),
            )
            pass_button.pack(side="left", padx=5, pady=5)

            # 불합격 = button
            fail_button = ctk.CTkButton(
                product_frame,
                text="fail",
                width=50,
                fg_color="#CC3D3D",
                hover_color="#960707",
                font=self.font_style,
                command=lambda PRODUCT_ID=product_id, frame=product_frame: self.fail_button_callback(
                    option, PRODUCT_ID, frame
                ),
            )
            fail_button.pack(side="left", pady=5)

            # 제품명 = label
            label = ctk.CTkLabel(product_frame, text=product_name, font=self.font_style)
            label.pack(side="left", padx=5, pady=5, anchor="center")

    def retouch_product_image(
        self,
        naver_coupang: str,
        amazon_iherb: str,
        id: str,
        image,
        filename: str,
        num: int,
    ):
        folder_path = os.path.join(f"./{amazon_iherb}/{id}", f"{naver_coupang}")
        os.makedirs(folder_path, exist_ok=True)

        if naver_coupang == "naver":
            new_image = ImageOps.pad(
                image,
                (self.naver_image_size - 40, self.naver_image_size - 40),
                color="white",
            )
        else:
            new_image = ImageOps.pad(
                image,
                (self.coupang_image_size - 40, self.coupang_image_size - 40),
                color="white",
            )
        border_thickness = 20
        image_with_border = ImageOps.expand(
            new_image, border=border_thickness, fill="white"
        )

        image_with_border.save(f"./{amazon_iherb}/{id}/{naver_coupang}/{filename}")

        print(f"{num + 1} => retouch complete!")

    def create_thumbnail_image(self, naver_coupang: str, amazon_iherb: str, id: str):
        image = Image.open(f"./{amazon_iherb}/{id}/{naver_coupang}/image1.jpg")

        if naver_coupang == "naver":
            new_image = ImageOps.pad(
                image,
                (self.naver_thumbnail_size - 40, self.naver_thumbnail_size - 40),
                color="white",
            )
        else:
            new_image = ImageOps.pad(
                image,
                (self.coupangr_thumbnail_size - 40, self.coupangr_thumbnail_size - 40),
                color="white",
            )
        border_thickness = 10
        image_with_border = ImageOps.expand(
            new_image, border=border_thickness, fill="white"
        )
        border_thickness = 10
        border_color = self.manageVaribles.get_thumbnail_border_color()
        thumbnail = ImageOps.expand(
            image_with_border, border=border_thickness, fill=border_color
        )

        thumbnail.save(f"./{amazon_iherb}/{id}/{naver_coupang}/thumbnail.jpg")

        print(f"{id} => thumbnail created!")

    def amazon_ocr(self, asin_code: str):
        time.sleep(1)

        num = 0
        while os.path.exists(f"./amazon/{asin_code}/image{num + 1}.jpg"):
            ocr_text: str = pytesseract.image_to_string(
                Image.open(f"./amazon/{asin_code}/image{num + 1}.jpg"), lang="eng"
            )

            suspicious_ingredients = self.execute_ocr(ocr_text)

            num += 1

        suspicious_ingredients_string = "\n".join(suspicious_ingredients)
        self.manageVaribles.update_suspicious_ingredients(suspicious_ingredients_string)

    def iherb_ocr(self, url: str):
        fake_header = self.manageVaribles.get_fake_header()
        response = requests.get(url, headers=fake_header)
        soup = BeautifulSoup(response.content, "html.parser")

        container = soup.find("div", {"class": "supplement-facts-container"})
        supplement_facts: str = container.get_text(strip=True)
        supplement_facts = supplement_facts.replace(
            "영양 성분 정보", "영양 성분 정보\n"
        )

        container = soup.find("div", {"class": "prodOverviewIngred"})
        other_ingredients: str = container.get_text(strip=True)
        other_ingredients = other_ingredients.replace("주요 성분", "주요 성분\n")
        other_ingredients = other_ingredients.replace("기타 성분", "\n기타 성분\n")

        if "무함유." in other_ingredients:
            other_ingredients = "무함유."
        elif "이 제품은" in other_ingredients:
            other_ingredients = other_ingredients.split("이 제품은")[0]
        elif "이 제품에는" in other_ingredients:
            other_ingredients = other_ingredients.split("이 제품에는")[0]

        translator = Translator()
        time.sleep(0.5)

        match = re.search(r"%하루 영양소 기준치(.*?)$", supplement_facts, re.DOTALL)
        if match:
            extracted_text = match.group(1).strip()
        supplement_facts: str = extracted_text

        text_to_translate: str = supplement_facts + " " + other_ingredients

        translated_text: str = translator.translate(text_to_translate, dest="en").text
        translated_text = (
            translated_text.replace("*", "")
            .replace("The standard value per day is not set.", ", ")
            .replace("Nothing.", "")
        )

        suspicious_ingredients = self.execute_ocr(translated_text)

        suspicious_ingredients_string = "\n".join(suspicious_ingredients)
        self.manageVaribles.update_suspicious_ingredients(suspicious_ingredients_string)

    def execute_ocr(self, ocr_text: str) -> List[str]:
        suspicious_ingredients = []

        with open("./2022.10.16.txt", "r") as f:
            word_list = [line.strip() for line in f.readlines() if line.strip()]

        found = False  # flag

        for word in word_list:
            if word.lower() in ocr_text.lower():
                ocr_words = ocr_text.split()

                for ocr_word in ocr_words:
                    if word.lower() in ocr_word.lower():
                        ocr_word = (
                            ocr_word.replace("(", "")
                            .replace(")", "")
                            .replace(".", "")
                            .replace(",", "")
                            .replace(":", "")
                        )
                        warning_message = f"[ {ocr_word} ] => {word}"
                        warning_message = warning_message.replace("?", "").replace(
                            "_", ""
                        )

                        if warning_message not in suspicious_ingredients:
                            suspicious_ingredients.append(warning_message)

                        found = True

                if not found:
                    pass

        return suspicious_ingredients

    def id_button_callback(
        self,
        amazon_iherb: str,
        ID: str,
    ):
        self.display_image_on_canvas(amazon_iherb, ID)
        self.display_suspicious_ingredients_on_textbox(amazon_iherb, ID)
        self.code_lable.configure(text=ID)
        self.display_warning(amazon_iherb, ID)

    def display_image_on_canvas(self, amazon_iherb: str, ID: str):
        image_path = f"./{amazon_iherb}/{ID}/image1.jpg"
        image = Image.open(image_path)
        new_image = ImageOps.pad(image, (180, 180), color="white")
        border_thickness = 10
        image_with_border = ImageOps.expand(
            new_image, border=border_thickness, fill="white"
        )
        photo = ImageTk.PhotoImage(image_with_border)
        self.product_image_canvas.create_image(100, 100, anchor="center", image=photo)
        self.product_image_canvas.image = photo

    def display_suspicious_ingredients_on_textbox(self, amazon_iherb: str, ID: str):
        value = self.excelCRUD.get_suspicious_ingredients(f"{amazon_iherb}", ID)

        self.suspicious_ingredients_textbox.delete("0.0", ctk.END)
        self.suspicious_ingredients_textbox.insert("0.0", value)

    def display_warning(self, amazon_iherb: str, ID: str):
        value = self.excelCRUD.get_suspicious_ingredients(f"{amazon_iherb}", ID)
        if value == "":
            self.warning_lable.configure(text="금지성분이 발견되지 않았습니다.")
        else:
            self.warning_lable.configure(text="의심되는 성분이 존재합니다!")

    def pass_button_callback(self, ID: str, frame):
        self.logger(f"'{ID}'는 합격입니다.")
        frame.configure(fg_color="#217346")

    def fail_button_callback(self, amazon_iherb: str, ID: str, frame):
        self.logger(f"'{ID}'는 불합격입니다.")
        frame.destroy()

        self.excelCRUD.delete_row(f"{amazon_iherb}", ID)

        folder_path = f"./{amazon_iherb}/{ID}"
        if os.path.exists(folder_path):
            shutil.rmtree(folder_path)


class ManageVariables:
    def __init__(self):
        self._amazon_iherb_option = "amazon"
        self._thumbnail_border_color = "white"
        self._amazon_urls: List[str] = []
        self._iherb_urls: List[str] = []
        self._suspicious_ingredients = ""
        self._fake_header = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36"
        }

    def update_amazon_iherb_option(self, value: str):
        if value.strip() == "아마존":
            self._amazon_iherb_option = "amazon"
        else:
            self._amazon_iherb_option = "iherb"
        print("amazon/iherb =>", self._amazon_iherb_option)

    def get_amazon_iherb_option(self) -> str:
        return self._amazon_iherb_option

    def update_thumbnail_border_color(self, value: str):
        self._thumbnail_border_color = value.strip()
        print("thumbnail border color =>", self._thumbnail_border_color)

    def get_thumbnail_border_color(self) -> str:
        return self._thumbnail_border_color

    def append_url(self, url: str):
        if self._amazon_iherb_option == "amazon":
            self._amazon_urls.append(url)
        else:
            self._iherb_urls.append(url)

    def remove_url(self, url: str):
        if self._amazon_iherb_option == "amazon":
            self._amazon_urls.remove(url)
        else:
            self._iherb_urls.remove(url)

    def get_urls(self) -> List[str]:
        if self._amazon_iherb_option == "amazon":
            return self._amazon_urls
        else:
            return self._iherb_urls

    def update_suspicious_ingredients(self, value: str):
        self._suspicious_ingredients = value
        print("suspicious ingredients =>", self._suspicious_ingredients)

    def get_suspicious_ingredients(self) -> str:
        return self._suspicious_ingredients

    def get_fake_header(self):
        return self._fake_header


class ExcelCRUD:
    def __init__(self):
        self._excel_file = openpyxl.load_workbook(
            "products.xlsx", data_only=True
        )  # from app.py

    def get_products_urls(self, option: str) -> List[str]:
        sheet = self._excel_file[option]

        products_urls = []
        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                products_urls.append(cell.value)
        return products_urls

    def update_row(
        self,
        option: str,
        id: str,
        product_name: str,
        suspicious_ingredients_string: str,
        url: str,
    ):
        row_values = [id, product_name, suspicious_ingredients_string, url]

        sheet = self._excel_file[option]
        sheet.append(row_values)
        self._excel_file.save("products.xlsx")

    def get_suspicious_ingredients(self, option: str, asin_code: str) -> str:
        sheet = self._excel_file[option]

        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=3):
            if row[0].value == asin_code:
                return row[2].value

        return ""

    def delete_row(self, option: str, asin_code: str) -> None:
        sheet = self._excel_file[option]

        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
            if row[0].value == asin_code:
                sheet.delete_rows(row[0].row)
                break

        self._excel_file.save("products.xlsx")
