import openpyxl
from typing import List


class ExcelCRUD:
    def __init__(self):
        self._excel_file = openpyxl.load_workbook(
            "products.xlsx", data_only=True
        )  # from app.py

    def get_all_product_url(self, option: str) -> List[str]:
        sheet = self._excel_file[option]

        products_urls = []
        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                products_urls.append(cell.value)
        return products_urls

    def get_product_url(self, option: str, id: str) -> str:
        sheet = self._excel_file[option]

        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=4):
            if row[0].value == id:
                return row[3].value

    def update_row(
        self,
        option: str,
        id: str,
        product_name: str,
        suspicious_ingredients_string: str,
        url: str,
    ) -> None:
        row_values = [id, product_name, suspicious_ingredients_string, url]

        sheet = self._excel_file[option]
        sheet.append(row_values)
        self._excel_file.save("products.xlsx")

    def get_suspicious_ingredients(self, option: str, id: str) -> str:
        sheet = self._excel_file[option]

        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=3):
            if row[0].value == id:
                return row[2].value

        return ""

    def delete_row(self, option: str, id: str) -> None:
        sheet = self._excel_file[option]

        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
            if row[0].value == id:
                sheet.delete_rows(row[0].row)
                break

        self._excel_file.save("products.xlsx")

    def check_product_exists(self, option: str, id: str) -> bool:
        sheet = self._excel_file[option]

        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
            if row[0].value == id:
                return True
        return False
